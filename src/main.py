#import src.dom.minidom

from xml.dom import minidom
from openpyxl import load_workbook, workbook
#from openpyxl import workbook

wb = load_workbook(filename="./excel/44-10_2019!.xlsx", data_only=True)
ws = wb.active

names = ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=1, values_only=True)
quants = ws.iter_rows(min_row=1, min_col=2, max_row=ws.max_row, max_col=2, values_only=True)
summs = ws.iter_rows(min_row=1, min_col=3, max_row=ws.max_row, max_col=3, values_only=True)
wb.close()

th_names = [item for sublist in names for item in sublist]
th_quants = [item for sublist in quants for item in sublist]
th_summs = [item for sublist in summs for item in sublist]

new1_names = []
new1_quants = []
new1_summs = []

