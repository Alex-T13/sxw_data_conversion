from pathlib import Path
from xml.dom import minidom
from openpyxl import load_workbook


_this_dir = Path(__file__).parent.resolve()
file_name = "Тек.ремонт ЗАВОД 25-07.03-22999,54.xlsx"
path = f"{_this_dir}/reports2021/{file_name}"

wb = load_workbook(filename=path, data_only=True)
ws = wb.active


names = ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=1, values_only=True)
quants = ws.iter_rows(min_row=1, min_col=2, max_row=ws.max_row, max_col=2, values_only=True)
summs = ws.iter_rows(min_row=1, min_col=3, max_row=ws.max_row, max_col=3, values_only=True)
wb.close()
print(quants)


th_names = [item for sublist in names for item in sublist]
th_quants = [item for sublist in quants for item in sublist]
th_summs = [item for sublist in summs for item in sublist]

print(th_summs)
# new1_names = []
# new1_quants = []
# new1_summs = []

# i = 0
# while i < len(th_names):
#     new1_names.append(th_names[i])
#     new1_quants.append(th_quants[i])
#     new1_summs.append(th_summs[i])
#     i += 1
# print(new1_summs)
# th_names = [item for sublist in names for item in sublist]
new_names = list(map(str, th_names))
# th_quants = [item for sublist in quants for item in sublist]
new_quants = list(map(float, th_quants))
# th_summs = [item for sublist in summs for item in sublist]
new_summs = list(map(float, th_summs))

print(new_summs)


# colA = ws[1]
# print(f"colA: {colA}")

# ws.iter_cols(min_col=1, max_col=1, min_row=1, max_row=ws.max_row, values_only=True)

for row in ws.iter_rows(min_col=3, max_col=3, max_row=ws.max_row, values_only=True):
    print(f"c: {row}")

# ns = list(map(float, ws.iter_cols(min_col=3, max_col=3, max_row=ws.max_row, values_only=True)))
# print(ns)

li = list(row for row in ws.iter_cols(min_col=3, max_col=3, max_row=ws.max_row, values_only=True))
print(f"li: {li}")

# c = [c * 3 for c in 'list']
th_n = [item for sublist in ws.iter_rows(min_col=3, max_col=3, max_row=ws.max_row, values_only=True) for item in sublist]
print(f"th_n: {th_n}")

r = 0
v = 0
for row in ws.values:
    r += 1
    v = 0
    for value in row:
        if v % 3:
            print(f"value: {value}")
        v += 1


# class Mat(models.Model):
#     name = models.CharField(max_length=255, db_index=True, verbose_name='Наименование')
#     quantity = models.DecimalField(max_digits=12, decimal_places=5, verbose_name='Количество')




# создаём объект
doc = minidom.Document()

# корневой тег root
root = doc.createElement('root')
doc.appendChild(root)

root1 = doc.createElement('root1')
root.appendChild(root1)

root2 = doc.createElement('root2')
root1.appendChild(root2)

#root3 = doc.createElement('root3')
# root2.appendChild(root3)

#root4 = doc.createElement('root4')
# root3.appendChild(root4)

#root5 = doc.createElement('root5')
# root4.appendChild(root5)

i = 0
while i < len(new_names):
    # print(user[i])
    if i < 9:
        obosn = ('СЦЕН-МТ-00'+str(i+1))
    elif i < 99:
        obosn = ('СЦЕН-МТ-0'+str(i+1))
    else:
        obosn = ('СЦЕН-МТ-'+str(i+1))

    mat = round(new_summs[i] / new_quants[i], 5)
    tr = round(mat * 0.096, 2)
    cena_0 = round(mat + tr, 2)

    # первый блок rascenka
    rascenka = doc.createElement('rascenka')
    rascenka.setAttribute('npp', str(i+1))  # 1
    rascenka.setAttribute('obosn', obosn)  # СЦЕН-МТ-001
    rascenka.setAttribute('naim', new_names[i])  # Сайдинг Extra color
    rascenka.setAttribute('idGrw', '7')
    rascenka.setAttribute('klv', str(new_quants[i]))  # klv
    rascenka.setAttribute('ed_izm', 'ШТ')
    rascenka.setAttribute('vid_ohr_pp', '12')
    rascenka.setAttribute('uniq_lab_ohr_pp', '5.1')
    rascenka.setAttribute('tip', '101')
    rascenka.setAttribute('cena_0', str(cena_0))  # 3.29

    # nabor_kf
    nabor_kf = doc.createElement('nabor_kf')
    rascenka.appendChild(nabor_kf)

    # koef_1
    koef_1 = doc.createElement('koef_1')
    koef_1.setAttribute('naim', '1-й коэффициент к расценке (справочно)')
    koef_1.setAttribute('mat', '1')
    koef_1.setAttribute('tr', '1')
    nabor_kf.appendChild(koef_1)

    # koef_2
    koef_2 = doc.createElement('koef_2')
    koef_2.setAttribute('naim', '2-й коэффициент к расценке (справочно)')
    koef_2.setAttribute('mat', '1')
    koef_2.setAttribute('tr', '1')
    nabor_kf.appendChild(koef_2)

    # koef_3
    koef_3 = doc.createElement('koef_3')
    koef_3.setAttribute('naim', '1-й коэффициент к расценке (справочно)')
    koef_3.setAttribute('mat', '1')
    koef_3.setAttribute('tr', '1')
    nabor_kf.appendChild(koef_3)

    # cena
    cena = doc.createElement('cena')
    cena.setAttribute('mat', str(mat))
    cena.setAttribute('tr', str(tr))
    cena.setAttribute('pryam', str(cena_0))
    cena.setAttribute('prc_ohr', '71.02')
    cena.setAttribute('prc_pp', '47.58')
    rascenka.appendChild(cena)

    # stoimost
    stoimost = doc.createElement('stoimost')
    stoimost.setAttribute('mat', str(new_summs[i]))
    stoimost.setAttribute('tr', str(round(tr * new_quants[i], 2)))
    stoimost.setAttribute('pryam', str(round(tr * new_quants[i] + new_summs[i], 2)))
    stoimost.setAttribute('ohr', '0')
    stoimost.setAttribute('pp', '0')
    rascenka.appendChild(stoimost)

    # materialy
    materialy = doc.createElement('materialy')
    rascenka.appendChild(materialy)

    # resurs
    resurs = doc.createElement('resurs')
    resurs.setAttribute('obosn', obosn)  # СЦЕН-МТ-001
    resurs.setAttribute('kodcic', '')
    resurs.setAttribute('naim', new_names[i])  # Сайдинг Extra color
    resurs.setAttribute('ed_izm', 'ШТ')
    resurs.setAttribute('norma', '1')  # klv
    resurs.setAttribute('klv', str(new_quants[i]))
    resurs.setAttribute('cena_0', str(mat))
    resurs.setAttribute('stm_0', str(round(mat * new_quants[i], 2)))
    resurs.setAttribute('tr', str(tr))
    resurs.setAttribute('tr_rub', '0')
    resurs.setAttribute('cena_tr', '9,6')
    materialy.appendChild(resurs)

    root2.appendChild(rascenka)

    i += 1

xml_str = doc.toprettyxml(indent="  ")
file_name_xml = path[:-4] + "xml"
with open(f"{file_name_xml}", "w") as f:
    f.write(xml_str)


""" 
<rascenka npp="1" obosn="СЦЕН-МТ-001" naim="Сайдинг Extra color могано 3660мм s-0.85m2/шт" idGrw="7" klv="15" ed_izm="ШТ" vid_ohr_pp="12" uniq_lab_ohr_pp="5.1" tip="101" cena_0="3.29">
      <nabor_kf>
       <koef naim="1-й коэффициент к расценке (справочно)" mat="1" tr="1"/>
       <koef naim="2-й коэффициент к расценке (справочно)" mat="1" tr="1"/>
       <koef naim="3-й коэффициент к расценке (справочно)" mat="1" tr="1"/>
       <koef_total_ohr_pp naim="Коэффициент по группе работ (5.1)" mat="1" tr="1" ohr="1" pp="1"/>
      </nabor_kf>
      <cena mat="3" tr="0.29" pryam="3.29" prc_ohr="71.02" prc_pp="47.58"/>
      <stoimost mat="45" tr="4.35" pryam="49.35" ohr="0" pp="0"/>
      <materialy>
       <resurs obosn="СЦЕН-МТ-001" kodcic="" naim="Сайдинг Extra color могано 3660мм s-0.85m2/шт" ed_izm="ШТ" norma="1" klv="15" cena_0="3" stm_0="45" tr="0.29" tr_rub="0" cena_tr="9.6"/>
      </materialy>
     </rascenka> 
"""
