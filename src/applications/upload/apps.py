from django.apps import AppConfig
from openpyxl import load_workbook

from project.settings import DIR_PROJECT


class UploadConfig(AppConfig):
    label = 'upload'
    name = f"applications.{label}"


def handle_uploaded_file(file):
    # if file.size()
    # wb = load_workbook(filename=file, data_only=True)
    # ws = wb.active
    # print(ws)
    new_file = f'{DIR_PROJECT}/temporary_files/file.xlsx'
    with open(new_file, 'wb') as destination:
        for chunk in file.chunks():
            destination.write(chunk)

    wb = load_workbook(filename=new_file, data_only=True)
    ws = wb.active
