from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename="10_08-2017.xlsx", data_only=True)
ws = wb.active

names = ws.iter_rows(min_row=1, min_col=2, max_row=ws.max_row, max_col=2, values_only=True)
quants = ws.iter_rows(min_row=1, min_col=4, max_row=ws.max_row, max_col=4, values_only=True)
summs = ws.iter_rows(min_row=1, min_col=3, max_row=ws.max_row, max_col=3, values_only=True)
wb.close()

print(type(summs))

th_names = [item for sublist in names for item in sublist]
new_names = list(map(str, th_names))
th_quants = [item for sublist in quants for item in sublist]
new_quants = list(map(float, th_quants))
th_summs = [item for sublist in summs for item in sublist]
#th_summs = [summs]

print(type(th_summs))

new_summs = list(map(float, th_summs))

print(type(new_summs))

#print(names, len(names))
#print(quants, len(quants))
#print(summs, len(summs))
#print(names[7])
#new_quants = list(map(int, quants))
#new_summs = list(map(float, summs))
print(new_summs, len(new_summs))

i = 0
while i < 10: #len(names):
    #th_quants = str(quants[i])
    #th_summs = str(summs[i])
    #new_list = list(map(int, old_list))
    #new_quants = map(float, th_quants)
    #new_summs = map(float, th_summs)
    #new_quants = int(th_quants)
    #new_summs = int(th_summs)
    #mat = new_summs[i] / new_quants[i]
    #print(mat)
    #th_summs = summs[i]
    #print(th_summs, len(th_summs))
    #print(summs[i])
    i += 1
